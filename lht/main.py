import openpyxl
import requests

def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))  # 确保每一行都是列表形式，以便修改
    return data

def call_chatgpt(input_data, prompt):
    url = 'YOUR_CHATGPT_OPENAPI_URL'
    headers = {
        'Authorization': 'Bearer YOUR_ACCESS_TOKEN',  # 替换成你的Authorization header
    }
    payload = {
        'input': input_data,
        'prompt': prompt,
    }
    response = requests.post(url, json=payload, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        print("Failed to call ChatGPT OpenAPI:", response.status_code)
        return None

def write_to_excel(file_path, data):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    for i, row in enumerate(data, start=1):
        if len(row) > 2 and row[2] is not None:  # 确保输出数据存在
            ws.cell(row=i, column=3, value=row[2])  # 将ChatGPT的输出写入第三列
    wb.save(file_path)

def main():
    excel_file = 'YOUR_EXCEL_FILE.xlsx'  # 替换成你的Excel文件路径
    excel_data = read_excel(excel_file)

    for row_index, row in enumerate(excel_data):
        if len(row) > 1:  # 确保每行至少有两个元素（输入和提示）
            input_data, prompt = row[0], row[1]
            chatgpt_result = call_chatgpt(input_data, prompt)
            if chatgpt_result and 'output' in chatgpt_result:
                excel_data[row_index].append(chatgpt_result['output'])  # 将结果添加为第三列

    write_to_excel(excel_file, excel_data)

if __name__ == "__main__":
    main()
